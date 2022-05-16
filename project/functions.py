import linecache

import jira
import openpyxl

"""
:keyword


"""


# Decorator for calculating consume time.
def timer(func):
    import functools
    from time import time

    @functools.wraps(func)
    def func_wrapper(*args, **kwargs):
        time_start = time()
        result = func(*args, **kwargs)
        time_spend = time() - time_start
        print('⏱️ FUNCTION: {} costs: {:.3f}s'.format(func.__name__, time_spend))
        return result

    return func_wrapper


# Create Jira data.
def create_jira():
    """
    Create jira using jira library.
    :return: jira_authorized
    """
    usr = linecache.getline('/Users/Elton/Code-stuff/userinfo.txt', 1).strip()
    pwd = linecache.getline('/Users/Elton/Code-stuff/userinfo.txt', 2).strip()

    jira_server = "https://help.bytedance.net"
    jira_authorized = jira.JIRA(jira_server, basic_auth=(usr, pwd))
    return jira_authorized


# Check JQL command
def check_jql(jira_authorized, jql):
    try:
        result = jira_authorized.search_issues(jql_str=jql,
                                               startAt=0,
                                               maxResults=1,
                                               json_result=True)
        total_issues = result["total"]
        print(f"⚠️ This JQL contains {total_issues} issues.")
        return total_issues
    except:
        return False


# Search JQL in jira, return single issues.
def search_jql_return_first(jira_authorized, jql):
    try:
        result = jira_authorized.search_issues(jql_str=jql,
                                               startAt=0,
                                               maxResults=1,
                                               json_result=True)
        return result
    except:
        return False


# Search JQL in jira, return multiple issues.
@timer
def search_jql_return_all(jira_authorized, jql, total_issues):
    num_max_results = 1000  # Jira only allowed download 1000 issues one time.

    from math import ceil
    query_times = ceil(total_issues / num_max_results)
    print(f"⚠️ All the issues need to download {query_times} times")

    result_all = []
    for i in range(query_times):
        start_at = num_max_results * i
        result_batch = jira_authorized.search_issues(jql_str=jql,
                                                     startAt=start_at,
                                                     maxResults=num_max_results,
                                                     json_result=True,
                                                     expand="changelog")

        result_all.append(result_batch)
    return result_all


# Search detail info from results,
def pick_field_from_issues(all_issues):
    fields = []

    for issue in all_issues:
        num_issues_batch = len(issue["issues"])
        print(f"⚠️ Total issues in this batch: {num_issues_batch}")

        for i in range(num_issues_batch):

            # Issue Type
            try:
                issue_type = issue["issues"][i]["fields"]["issuetype"]["name"]
            except:
                issue_type = "notFound"
            # Issue Key
            try:
                issue_key = issue["issues"][i]["key"]
            except:
                issue_key = "notFound"
            # Summary
            try:
                summary = issue["issues"][i]["fields"]["summary"]
            except:
                summary = "notFound"
            # Assignee
            try:
                issue_assignee = issue["issues"][i]["fields"]["assignee"]["displayName"]
            except:
                issue_assignee = "notFound"
            # Assignee Email
            try:
                issue_assignee_email = issue["issues"][i]["fields"]["assignee"]["emailAddress"]
            except:
                issue_assignee_email = "notFound"
            # Create Time
            try:
                created_date_time = issue["issues"][i]["fields"]["created"]
                created_date_time = created_date_time.split('T')[0] + ' ' + \
                                    created_date_time.split('T')[1].split('000')[0]
                created_date = created_date_time.split(" ")[0]
                created_time = created_date_time.split(" ")[1]
            except:
                created_date_time = "notFound"
                created_date = "notFound"
                created_time = "notFound"
            # SLA
            try:
                remaining_time = issue["issues"][i]["fields"]["customfield_10104"]["ongoingCycle"]["remainingTime"][
                    "friendly"]
            except:
                remaining_time = "notFound"
            # Issue Link
            try:
                issue_link = "https://help.bytedance.net/browse/" + issue["issues"][i]["key"]
            except:
                issue_link = "notFound"

            field = [summary,
                     issue_assignee,
                     issue_assignee_email,
                     created_date,
                     created_time,
                     remaining_time,
                     issue_link]

            fields.append(field)
    return fields


# Get summary index in lists
def get_summary_index(fields):
    index_num = 0
    try:
        for field in fields:
            for element in field:
                if "|" in element:
                    index_num = field.index(element)
                continue
    except:
        for field in fields:
            for element in field:
                if "【" in element:
                    index_num = field.index(element)
                continue

    return index_num


# Replace the data type
def replace_alarm_type_to_filter(string):
    """
    replace the alarm_type from Feishu to Filter.
    Left: Feishu alarm, Right: Filter alarm.
    """

    string = string.replace("飞书会议室离线", "主机离线")
    string = string.replace("会议室检测到高 CPU 使用率", "主机CPU大于阀值")
    string = string.replace("控制器电量低", "控制器电池电量低于20%")
    string = string.replace("控制器已断开连接", "控制器离线")
    string = string.replace("签到板电池电量低", "签到板电池电量少于20%")
    string = string.replace("签到板离线", "签到板离线")
    string = string.replace("麦克风已断开连接", "主机麦克风断开连接")
    string = string.replace("摄像头已断开连接", "主机摄像头断开连接")
    string = string.replace("扬声器已断开连接", "主机扬声器断开连接")
    string = string.replace("飞书会议室离线", "主机离线")

    return string


# Only processing the summary string.
def processing_summary(string):
    processed_summary = []

    # Distinguish two types of issue, and processing data.
    if "|" in string:
        type = "过滤平台"
        city_site_room = string.split("|")[1]
        city = city_site_room.split("-")[0]
        site = city_site_room.split("-")[1]
        room = city_site_room.split(")")[1].strip("-")
        capacity = "Null"
        alarm_type = string.split("|")[2]

        processed_summary = [type, city, site, room, capacity, alarm_type]


    elif "【" in string:
        type = "飞书告警"
        city_site_room = string.split("】 ")[1]
        city = city_site_room.split("-")[0]
        site = city_site_room.split("-")[1]
        room = city_site_room.split(")")[1].strip("-").split("(")[0]
        capacity = city_site_room.split("(")[2].split(")")[0]
        alarm_type = city_site_room.split(")")[2].strip(" ")
        alarm_type = replace_alarm_type_to_filter(alarm_type)

        processed_summary = [type, city, site, room, capacity, alarm_type]

    return processed_summary


# Process the Summary to human-readable.
@timer
def process_feishu_summary(fields):
    """
    :param fields: Read a list, each element is a list that contains summary etcetera.
    :return: Two lists, one is the normal data, and the other is the exceptional data.
    """

    # Find where the summary is.
    summary_index = get_summary_index(fields)

    # Index numbers
    num_of_index = 0

    # Create an empty ist to  storage processed issues.
    results = []

    # Create an empty list to storage exceptional issues.
    exception = []

    for field in fields:
        summary_original = field[summary_index]

        # Human-readable.
        num_of_index = num_of_index + 1

        # Remove ZOOM
        if "Zoom" in summary_original:
            exception.append(field)
            continue
        # Remove Alert
        if "Alert" in summary_original:
            exception.append(field)
            continue
        # Remove Update
        if "Update" in summary_original:
            exception.append(field)
            continue
        # Remove XSK
        if "XSK" in summary_original:
            exception.append(field)
            continue
        # Remove Test
        if "Test" in summary_original:
            exception.append(field)
            continue

        # return list of processed summary
        processed_summary = processing_summary(summary_original)

        # Using Slice to insert processed summary to each field.
        field[1:1] = processed_summary

        results.append(field)

    return results, exception


# Create an Excel file to write data.
def create_and_write_excel(path, value):
    '''
    :param path: File path that created by write_to_excel function.
    :param value: A list in list form data.
    :return: A Excel file that only contains one sheet.
    '''
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "汇总"

    for row in range(len(value)):
        for column in range(len(value[0])):
            sheet.cell(row=row + 1, column=column + 1, value=str(value[row][column]))

    book.save(path)
    print("Creating & Writing Finished.")


# Append data write to existing Excel file .
def add_excel(path, value, sheet):
    """
    :param path: File path that created by write_to_excel function.
    :param value: A list in list form data.
    :param sheet: A sheet name that directly named in the function.
    :return: Adding the sheet to an existing Excel file.
    """
    wb = openpyxl.load_workbook(path)
    wb.create_sheet(sheet)
    ws = wb[sheet]
    for ss in value:
        ws.append(ss)
    wb.save(path)
    print("append Success.")


# Export the final file.
def write_to_excel(final, result, exception):
    """
    :param final: All the data returned from jira.
    :param result: All the normal data.
    :param exception: All the exceptional data.
    :return:
    """

    import datetime
    now_time = datetime.datetime.now().strftime('%Y-%m-%d %H时%M分%S秒')
    path = f"/Users/Elton/Downloads/20220510_告警数据/告警数据 {now_time}" + ".xlsx"

    create_and_write_excel(path, final)

    add_excel(path, result, "告警汇总")
    add_excel(path, exception, "例外情况")
    add_excel(path, [], "透视图")
